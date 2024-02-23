from openpyxl import load_workbook


class Workbook():
    def __init__(self) -> None:
        from io import BytesIO
        workbook_path = "resumo/static/resumo.xlsx"
        self.workbook = load_workbook(workbook_path)
        self.virtual_wb = BytesIO()

    def write_to_excel(self, spreadsheet, row, page, analysis) -> None:
        def str_to_float(string) -> float:
            return float(string.replace(".", "").replace(",", "."))

        try:
            spreadsheet.cell(row, 2).value = page.page_number
            spreadsheet.cell(row, 3).value = page.date
            spreadsheet.cell(row, 4).value = page.representative
            spreadsheet.cell(row, 5).value = page.buyer
            spreadsheet.cell(row, 6).value = page.buyer_cpf
            spreadsheet.cell(row, 7).value = page.serial_number
            spreadsheet.cell(row, 8).value = page.analysis_type
            spreadsheet.cell(row, 9).value = str_to_float(analysis["kg"])
            spreadsheet.cell(row, 10).value = str_to_float(analysis["pd"])
            spreadsheet.cell(row, 11).value = str_to_float(analysis["pt"])
            spreadsheet.cell(row, 12).value = str_to_float(analysis["rh"])
            spreadsheet.cell(row, 13).value = str_to_float(analysis["unitary_value"])
            spreadsheet.cell(row, 14).value = str_to_float(analysis["total_value"])
            spreadsheet.cell(row, 15).value = page.quotations[0]
            spreadsheet.cell(row, 16).value = page.quotations[1]
            spreadsheet.cell(row, 17).value = page.quotations[2]

        except TypeError as E:
            spreadsheet.cell(row, 1).value = "error"
            print(row, E)

        except Exception as E:
            print(row, E)

        finally:
            self.workbook.save(self.virtual_wb)

    def write(self, pdf, resumed: bool = False) -> None:
        def return_last_empty_row(spreadsheet) -> int:
            for row in spreadsheet.iter_rows(min_row=3, min_col=2, max_col=2):
                if row[0].value is None:
                    return row[0].row

        spreadsheet = self.workbook.active

        for page in pdf.pages:
            last_empty_row = return_last_empty_row(spreadsheet)

            if resumed:
                self.write_to_excel(spreadsheet, last_empty_row, page, page.analysis_average)

            if not resumed:
                for i, analysis in enumerate(page.analytics):
                    row = last_empty_row + i
                    self.write_to_excel(spreadsheet, row, page, analysis)

        return self.virtual_wb
