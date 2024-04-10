from io import BytesIO
from modules.utils import load_config, return_last_empty_row, str_to_float, cleanse_digits_list


class Pdf:
    def __init__(self, file: BytesIO, coordinates="br_coordinates") -> None:
        import fitz
        self.pages = list()
        for page in fitz.open(stream=file): self.pages.append(PdfPage(page, coordinates))


class PdfPage:
    def __init__(self, properties, coordinates) -> None:
        # top left coordinates (x0, y0) and bottom right coordinates (x1, y1)
        # this is measured in points (1 inch = 72 points), you need to get this manually
        self.coordinates = load_config("coordinates", coordinates)
        self.properties = properties
        self.page_number = self.properties.number + 1

        self.date = self.ocr("date")
        self.representative = self.ocr("representative")
        self.buyer = self.ocr("buyer")
        self.analysis_type = int(self.ocr("analysis_type"))
        self.buyer_cpf = self.ocr("buyer_cpf")
        self.buyer_rg = self.ocr("buyer_rg")
        self.serial_number = self.ocr("serial_number")
        self.quotations = cleanse_digits_list(self.ocr("quotations").split("\n"))
        self.analysis = cleanse_digits_list(self.ocr("analysis").split("\n"))
        self.analysis_average = self.ocr("analysis_average").split("\n")
        self.avg_kg_price = self.ocr("avg_kg_price")
        self.total_value = self.ocr("total_value")

        self.analysis_lines_treatment()
        self.analysis_average_treatment()

    def ocr(self, c):
        return self.properties.get_text("text", self.coordinates[c]).strip()

    def analysis_lines_treatment(self) -> list:
        from math import ceil
        unformatted_list = self.analysis
        col_count = 6
        formatted_list = list()
        list_lenght = int(ceil(len(unformatted_list)) / col_count)

        # groups the old list items together in new lists the size of column_count
        for i in range(list_lenght):
            c = (i * col_count)
            kg, pd, pt, rh, unitary_value, total_value = unformatted_list[c:c + col_count]
            formatted_list.append({"kg": kg, "pd": pd, "pt": pt, "rh": rh,
                                   "unitary_value": unitary_value, "total_value": total_value})

        self.analysis = formatted_list


    def analysis_average_treatment(self) -> dict:
        from string import ascii_letters

        analysis_dict = dict()

        analysis_dict["pd"] = str_to_float(self.analysis_average[0])
        analysis_dict["pt"] = str_to_float(self.analysis_average[1])
        analysis_dict["rh"] = str_to_float(self.analysis_average[2])
        analysis_dict["kg"] = str_to_float(self.analysis_average[3])
        analysis_dict["avg_kg_price"] = str_to_float(self.avg_kg_price)
        analysis_dict["total_value"] = str_to_float(self.total_value.strip(ascii_letters).strip("/\\"))

        self.analysis_average = analysis_dict


class Workbook():
    def initialize(self, write_type):
        from openpyxl import load_workbook

        working_type = load_config("types", str(write_type))
        self.workbook = load_workbook(working_type["wb_path"])
        self.workbook.active = 0
        self.virtual_wb = BytesIO()

        return eval(working_type["fun"])

    def write_line_by_line(self, spreadsheet, page, last_page, repeat_page_number = True, *args) -> None:
        try:
            analytics = page.analysis
            
            for analysis in analytics:
                iteration = [page.page_number if last_page != page.page_number or repeat_page_number else "", 
                             page.date, page.representative, page.buyer, page.buyer_cpf, page.serial_number,
                             int(page.analysis_type), analysis["kg"], analysis["pd"], analysis["pt"], analysis["rh"],
                             analysis["total_value"], page.quotations[0], page.quotations[1], page.quotations[2]]
                row = return_last_empty_row(spreadsheet)
                last_page = page.page_number

                for i, it in enumerate(iteration):
                    i = i + 2
                    spreadsheet.cell(row, i).value = it

        except TypeError as E:
            spreadsheet.cell(row, 1).value = "error"
            print(row, E)

        except Exception as E:
            print(E)

        finally:
            self.virtual_wb.seek(0)
            self.workbook.save(self.virtual_wb)
            return last_page

    def write_brief(self, spreadsheet, page, *args) -> None:
        try:
            analysis = page.analysis_average
            row = return_last_empty_row(spreadsheet)

            iteration = [page.page_number, page.date, page.representative, page.buyer, page.buyer_cpf,
                        page.serial_number, page.analysis_type, analysis["kg"], analysis["pd"], 
                        analysis["pt"], analysis["rh"], analysis["total_value"], page.quotations[0], 
                        page.quotations[1], page.quotations[2]]

            for i, it in enumerate(iteration):
                i = i + 2

                spreadsheet.cell(row, i).value = it

        except TypeError as E:
            spreadsheet.cell(row, 1).value = "error"
            print(row, E)

        except Exception as E:
            print(E)

        finally:
            self.virtual_wb.seek(0)
            self.workbook.save(self.virtual_wb)

    def write_general_brief(self, spreadsheet, page, *args) -> None:
        try:
            analysis = page.analysis_average
            row = return_last_empty_row(spreadsheet)

            iteration = [page.page_number, len(page.analysis), page.buyer[:page.buyer.find(" ")],
                         page.date[:page.date.find(" ")], page.date[page.date.find(" ")+1:],
                         analysis["kg"], analysis["total_value"], int(page.serial_number[-3:])]

            for i, it in enumerate(iteration):
                i = i + 2

                spreadsheet.cell(row, i).value = it

        except TypeError as E:
            spreadsheet.cell(row, 1).value = "error"
            print(row, E)

        except Exception as E:
            print(E)

        finally:
            self.virtual_wb.seek(0)
            self.workbook.save(self.virtual_wb)

    def write(self, pdf, write_type) -> None:
        last_page = 0

        fun = self.initialize(write_type)
        spreadsheet = self.workbook.active

        for page in pdf.pages:
            fun(spreadsheet, page, last_page)

        self.virtual_wb.seek(0)
        return self.virtual_wb
