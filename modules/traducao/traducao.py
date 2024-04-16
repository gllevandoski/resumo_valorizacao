from openpyxl import load_workbook
from io import BytesIO


class Translate:
    def __init__(self, brief_workbook) -> None:
        self.brief_workbook = load_workbook(brief_workbook)
        self.batch_spreadsheet = self.brief_workbook.active

        self.template_workbook = "modules/traducao/base.xlsx"

        self.analysis_list = list()
        self.workbooks = list()
        self.pdfs = list()

        self.read()
        self.generate_sheets()
        pdfs = self.generate_pdfs()
        self.merge_pdfs(pdfs)

    def read(self):
        cfg = load_config("traducao/config")

        # iter through the rows filled with content
        for row in self.batch_spreadsheet.iter_rows(min_row=3, min_col=2, values_only=True):
            if row[0] is None:
                break

            analytics = dict()

            analytics["kg"] = row[cfg["analytics_kg"]]
            analytics["pd"] = row[cfg["analytics_pd"]]
            analytics["pt"] = row[cfg["analytics_pt"]]
            analytics["rh"] = row[cfg["analytics_rh"]]
            analytics["total_value"] = row[cfg["analytics_total_value"]]

            self.analysis_list.append({
                "page_number": row[cfg["page_number"]],
                "representative": row[cfg["representative"]],
                "buyer": row[cfg["buyer"]],
                "buyer_cpf": row[cfg["buyer_cpf"]],
                "serial_number": row[cfg["serial_number"]],
                "analysis_type": row[cfg["analysis_type"]],
                "analytics": [analytics],
                "pd_price": row[cfg["pd_price"]],
                "pt_price": row[cfg["pt_price"]],
                "rh_price": row[cfg["rh_price"]]
            })

    def generate_sheets(self):
        def return_first_empty_row(spreadsheet):
            for row in spreadsheet.iter_rows(min_row=20, max_row=34, min_col=3, max_col=3):
                if row[0].value is None:
                    return row[0].row

        def format_info(info):
            formated = dict()

            for i in range(1, 200):
                for f in info:
                    if f["page_number"] == i:
                        if i in formated:
                            formated[i]["analytics"].append(f["analytics"][0])
                        else:
                            formated[i] = f

            return formated

        formated_info = format_info(self.analysis_list)

        for info in formated_info.values():
            template = load_workbook(self.template_workbook)
            main_spreadsheet = template["Média de Valorização"]
            calc_spreadsheet = template["Interna"]

            calibrations = load_config("calibration")[info["serial_number"]]

            calc_spreadsheet["d15"] = calibrations["pd"]
            calc_spreadsheet["d16"] = calibrations["pt"]
            calc_spreadsheet["d17"] = calibrations["rh"]

            main_spreadsheet["c4"] = info["representative"]
            main_spreadsheet["cf14"] = info["analysis_type"]
            main_spreadsheet["e4"] = info["buyer"]
            main_spreadsheet["e6"] = info["buyer_cpf"]
            main_spreadsheet["x9"] = info["pd_price"]
            main_spreadsheet["x10"] = info["pt_price"]
            main_spreadsheet["x11"] = info["rh_price"]
            main_spreadsheet["w14"] = info["serial_number"]

            for analysis in info["analytics"]:
                row = return_first_empty_row(main_spreadsheet)
                main_spreadsheet.cell(row, 3).value = analysis["kg"]
                main_spreadsheet.cell(row, 4).value = analysis["pd"]
                main_spreadsheet.cell(row, 5).value = analysis["pt"]
                main_spreadsheet.cell(row, 6).value = analysis["rh"]

            # saving the spreadsheet
            self.workbooks.append(template)

    def generate_pdfs(self):
        from win32com import client
        import pythoncom
        from uuid import uuid4 as uid
        pythoncom.CoInitialize()

        pdfs = list()
        path = fr"C:\Users\Home\Documents\Python\prog\resumo_valorizacao\.tmp\base.xlsx"

        for workbook in self.workbooks:
            workbook.save(path)
            with open(self.template_workbook):
                try:
                    excel = client.Dispatch("Excel.Application")
                    excel.Application.DisplayAlerts = False

                    wb = excel.Workbooks.OpenXML(path)
                    ws = wb.Worksheets[1]

                    ws.PageSetup.PrintArea = "B1:CB53"
                    uid_str = str(uid()).replace("-", "")
                    ws.ExportAsFixedFormat(0, fr"C:\Users\Home\Documents\Python\prog\resumo_valorizacao\.tmp\{uid_str}.pdf")
                    pdfs.append(fr"C:\Users\Home\Documents\Python\prog\resumo_valorizacao\.tmp\{uid_str}.pdf")

                finally:
                    wb.Close(False)
                    excel.Application.Quit()
                    excel = None
                    wb = None
                    ws = None

        pythoncom.CoUninitialize()
        return pdfs

    def merge_pdfs(self, pdfs):
        import fitz

        file = BytesIO()
        result = fitz.open()

        for pdf in pdfs:
            with fitz.open(pdf) as mfile:
                result.insert_pdf(mfile)

        result.save(file)

        self.merged_pdf = file


def load_config(filename):
    with open(f"modules/{filename}.json") as cfg_file:
        cfg = cfg_file.read()
    from json import loads
    return loads(cfg)
