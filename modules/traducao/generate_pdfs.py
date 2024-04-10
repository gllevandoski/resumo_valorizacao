from openpyxl import load_workbook # https://openpyxl.readthedocs.io/en/stable/
from win32com import client


def load_new_wb(filename):
    destination_folder = "sheets"
    workbook = load_workbook("assets/base.xlsx")
    workbook.save(f"{destination_folder}/{filename}.xlsx")
    return load_workbook(f"{destination_folder}/{filename}.xlsx")


def save_wb(workbook, filename):
    destination_folder = "sheets"
    workbook.save(f"{destination_folder}/{filename}.xlsx")


def read_from_excel():
    analysis_list = list()
    wb = load_workbook("assets/resumo.xlsx")
    spreadsheet = wb.active

    for row in spreadsheet.iter_rows(min_row=3, min_col=2, values_only = True):
        if row[0] is None:
            break

        analytics = dict()

        page_number = row[0]
        representative = row[2]
        buyer = row[3]
        buyer_cpf = row[4]
        serial_number = row[5]
        analysis_type = row[6]

        analytics["kg"] = row[7]
        analytics["pd"] = row[8]
        analytics["pt"] = row[9]
        analytics["rh"] = row[10]
        # analytics["total_value"] = row[11]

        pd_price = row[12]
        pt_price = row[13]
        rh_price = row[14]

        # pd_lease = row[15]
        # pt_lease = row[16]
        # rh_lease = row[17]

        analysis_list.append({
            "page_number": page_number,
            "representative": representative,
            "buyer": buyer,
            "buyer_cpf": buyer_cpf,
            "serial_number": serial_number,
            "analysis_type": analysis_type,
            "analytics": [analytics],
            "pd_price": pd_price,
            "pt_price": pt_price,
            "rh_price": rh_price,
            # "pd_lease": pd_lease,
            # "pt_lease": pt_lease,
            # "rh_lease": rh_lease
        })

    return analysis_list


def format_info(info):
    formated = dict()

    for i in range(1, 200):
        for f in info:
            if f["page_number"] == i:
                if i in formated:
                    formated[i]["analytics"].append(f["analytics"][0])
                else:
                    formated[i] = f

    return formated


def return_last_empty_row(spreadsheet):
    for row in spreadsheet.iter_rows(min_row=20, max_row=34, min_col=3, max_col=3):
        if row[0].value is None:
            return row[0].row


def generate_pdfs_from_xlsx():
    infor = read_from_excel()
    formated_info = format_info(infor)

    for info in formated_info.values():
        workbook = load_new_wb(info["page_number"])

        # # configuring spreadsheet
        # spreadsheet_int = workbook["Interna"]
        # spreadsheet_int["e8"] = info["pd_lease"]
        # spreadsheet_int["e9"] = info["pt_lease"]
        # spreadsheet_int["e10"] = info["rh_lease"]

        spreadsheet = workbook["Média de Valorização"]
        spreadsheet["c4"] = info["representative"]
        spreadsheet["cf14"] = info["analysis_type"]
        spreadsheet["e4"] = info["buyer"]
        spreadsheet["e6"] = info["buyer_cpf"]
        spreadsheet["x9"] = info["pd_price"]
        spreadsheet["x10"] = info["pt_price"]
        spreadsheet["x11"] = info["rh_price"]

        for analysis in info["analytics"]:
            row = return_last_empty_row(spreadsheet)
            spreadsheet.cell(row, 3).value = analysis["kg"]
            spreadsheet.cell(row, 4).value = analysis["pd"]
            spreadsheet.cell(row, 5).value = analysis["pt"]
            spreadsheet.cell(row, 6).value = analysis["rh"]

        # saving the spreadsheet
        save_wb(workbook, info["page_number"])

def save_to_pdf():
    from glob import glob

    files = glob(f"sheets/**/*.xlsx", recursive=True)

    for file in files:
        name = file.lstrip("sheets\\").rstrip(".xlsx")
        inp = f"C:/Users/Home/Documents/Python/prog/traduz_valorizacao/sheets/{name}.xlsx"
        out = f"C:/Users/Home/Documents/Python/prog/traduz_valorizacao/pdf/{name}.pdf"

        excel = client.Dispatch("Excel.Application")
        wb = excel.Workbooks.Open(inp)
        ws = wb.Worksheets[1]

        ws.PageSetup.PrintArea = "$B$1:$CB$53"
        ws.ExportAsFixedFormat(0, out)
        wb.Close(SaveChanges=1)

# generate_pdfs_from_xlsx()
save_to_pdf()
