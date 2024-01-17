from openpyxl import load_workbook


class Workbook():
    def __init__(self, workbook_path: str) -> None:
        self.workbook_path = workbook_path
        self.workbook = load_workbook(self.workbook_path)

    # def load_workbook(self, clone_name: str) -> workbook:
    #     try:
    #         if cloned_workbook := load_workbook(f"valorizacoes/{clone_name}.xlsx"):
    #             return cloned_workbook
    #     except FileNotFoundError:
    #         self.workbook.save(f"valorizacoes/{clone_name}.xlsx")
    #         cloned_workbook = load_workbook(f"valorizacoes/{clone_name}.xlsx")
    #         return cloned_workbook

    @staticmethod
    def str_to_float(string) -> float:
        return float(string.replace(".", "").replace(",", "."))

    def write(self, folders_list: list, resumed: bool = False):
        def return_last_empty_row(spreadsheet):
            for row in spreadsheet.iter_rows(min_row=3, min_col=2, max_col=2):
                if row[0].value is None:
                    return row[0].row

        spreadsheet = self.workbook.active

        for representative in folders_list.keys():
            for pdf in folders_list[representative]:
                for page in pdf.pages:
                    last_empty_row = return_last_empty_row(spreadsheet)

                    if resumed:
                        try:
                            spreadsheet.cell(last_empty_row, 2).value = page.properties.number + 1
                            spreadsheet.cell(last_empty_row, 3).value = page.date
                            spreadsheet.cell(last_empty_row, 4).value = page.representative
                            spreadsheet.cell(last_empty_row, 5).value = page.buyer
                            spreadsheet.cell(last_empty_row, 6).value = page.buyer_cpf
                            spreadsheet.cell(last_empty_row, 7).value = page.serial_number
                            spreadsheet.cell(last_empty_row, 8).value = page.analysis_type
                            spreadsheet.cell(last_empty_row, 9).value = self.str_to_float(page.analysis_average["kg"])
                            spreadsheet.cell(last_empty_row, 10).value = self.str_to_float(page.analysis_average["pd"])
                            spreadsheet.cell(last_empty_row, 11).value = self.str_to_float(page.analysis_average["pt"])
                            spreadsheet.cell(last_empty_row, 12).value = self.str_to_float(page.analysis_average["rh"])
                            spreadsheet.cell(last_empty_row, 14).value = self.str_to_float(page.analysis_average["total_value"])

                        except TypeError as E:
                            spreadsheet.cell(last_empty_row, 1).value = "error"
                            print(last_empty_row, E)

                        except Exception as E:
                            print(last_empty_row, E)
                            print(page.analysis_average)
                            print(pdf.pdf_path)

                        finally:
                            self.workbook.save(self.workbook_path)
                            break

                    if not resumed:
                        for i, analysis in enumerate(page.analytics):
                            row = last_empty_row + i

                            try:
                                spreadsheet.cell(row, 2).value = page.properties.number + 1
                                spreadsheet.cell(row, 3).value = page.date
                                spreadsheet.cell(row, 4).value = page.representative
                                spreadsheet.cell(row, 5).value = page.buyer
                                spreadsheet.cell(row, 6).value = page.buyer_cpf
                                spreadsheet.cell(row, 7).value = page.serial_number
                                spreadsheet.cell(row, 8).value = page.analysis_type
                                spreadsheet.cell(row, 9).value = self.str_to_float(analysis["kg"])
                                spreadsheet.cell(row, 10).value = self.str_to_float(analysis["pd"])
                                spreadsheet.cell(row, 11).value = self.str_to_float(analysis["pt"])
                                spreadsheet.cell(row, 12).value = self.str_to_float(analysis["rh"])
                                spreadsheet.cell(row, 13).value = self.str_to_float(analysis["unitary_value"])
                                spreadsheet.cell(row, 14).value = self.str_to_float(analysis["total_value"])

                            except TypeError as E:
                                spreadsheet.cell(row, 1).value = "error"
                                print(row, E)

                            finally:
                                self.workbook.save(self.workbook_path)


if __name__ == "__main__":
    import pdf

    pdfs = pdf.load()
    wb = Workbook("assets/resumo.xlsx")
    wb.write(pdfs, resumed=False)
    print("done")
