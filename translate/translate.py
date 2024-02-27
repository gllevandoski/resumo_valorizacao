from openpyxl import load_workbook


class Translate:
    def __init__(self, batch_workbook, template_workbook) -> None:
        self.batch_workbook = load_workbook(batch_workbook)
        self.batch_spreadsheet = self.batch_workbook.active

        self.template_workbook = template_workbook

        self.analysis_list = list()
        self.workbooks = list()
        self.pdfs = list()

        self.read()
        self.generate_sheets()
        self.generate_pdfs()

    @staticmethod
    def load_config(context):
        with open("translate/config.json") as cfg_file:
            cfg = cfg_file.read()

        from json import loads
        return loads(cfg)[context]

    def read(self):
        cfg = self.load_config("workbook")
        
        for row in self.batch_spreadsheet.iter_rows(min_row=3, min_col=2, values_only=True):
            if row[0] is None:
                break

            analytics = dict()

            analytics["kg"] = row[cfg["analytics_kg"]]
            analytics["pd"] = row[cfg["analytics_pd"]]
            analytics["pt"] = row[cfg["analytics_pt"]]
            analytics["rh"] = row[cfg["analytics_rh"]]
            analytics["total_value"] = row[cfg["analytics_total_value"]]

            self.analysis_list.append({
                "page_number": row[cfg["page_number"]],
                "representative": row[cfg["representative"]],
                "buyer": row[cfg["buyer"]],
                "buyer_cpf": row[cfg["buyer_cpf"]],
                "serial_number": row[cfg["serial_number"]],
                "analysis_type": row[cfg["analysis_type"]],
                "analytics": [analytics],
                "pd_price": row[cfg["pd_price"]],
                "pt_price": row[cfg["pt_price"]],
                "rh_price": row[cfg["rh_price"]]
            })

    def generate_sheets(self):
        def return_first_empty_row(spreadsheet):
            for row in spreadsheet.iter_rows(min_row=20, max_row=34, min_col=3, max_col=3):
                if row[0].value is None:
                    return row[0].row

        def format_info(info):
            formated = dict()

            for i in range(1, 200):
                for f in info:
                    if f["page_number"] == i:
                        if i in formated:
                            formated[i]["analytics"].append(f["analytics"][0])
                        else:
                            formated[i] = f

            return formated

        formated_info = format_info(self.analysis_list)

        for info in formated_info.values():
            template = load_workbook(self.template_workbook)
            main_spreadsheet = template["Média de Valorização"]

            main_spreadsheet["c4"] = info["representative"]
            main_spreadsheet["cf14"] = info["analysis_type"]
            main_spreadsheet["e4"] = info["buyer"]
            main_spreadsheet["e6"] = info["buyer_cpf"]
            main_spreadsheet["x9"] = info["pd_price"]
            main_spreadsheet["x10"] = info["pt_price"]
            main_spreadsheet["x11"] = info["rh_price"]

            for analysis in info["analytics"]:
                row = return_first_empty_row(main_spreadsheet)
                main_spreadsheet.cell(row, 3).value = analysis["kg"]
                main_spreadsheet.cell(row, 4).value = analysis["pd"]
                main_spreadsheet.cell(row, 5).value = analysis["pt"]
                main_spreadsheet.cell(row, 6).value = analysis["rh"]

            # saving the spreadsheet
            self.workbooks.append(template)

    def generate_pdfs(self):
        from win32com import client
        from io import BytesIO
        import pythoncom
        pythoncom.CoInitialize()

        for i, workbook in enumerate(self.workbooks):
            workbook.save(".tmp/tmpfile.xlsx")
            inp = open(".tmp/tmpfile.xlsx")

            try:
                excel = client.Dispatch("Excel.Application")

                wb = excel.Workbooks.OpenXML(r"C:\Users\Home\Desktop\resumo_valorizacao\.tmp\tmpfile.xlsx")
                ws = wb.Worksheets[1]

                ws.PageSetup.PrintArea = "B1:CB53"
                ws.ExportAsFixedFormat(0, fr"C:\Users\Home\Desktop\resumo_valorizacao\.tmp\{i}.pdf")
                wb.Close(True)
            finally:
                ws = None
                wb = None
                excel.Application.Quit()
                excel = None
                inp.close()
