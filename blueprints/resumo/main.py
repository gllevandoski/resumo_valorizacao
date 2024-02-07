from openpyxl import load_workbook


class Workbook():
    def __init__(self, workbook_template_path: str= "resumo.xlsx", output_name: str = "resumos.xlsx") -> None:
        self.workbook_template_path = f"assets/{workbook_template_path}"
        self.workbook_path = f"generated/{output_name}.xlsx"
        self.copy(self.workbook_template_path, self.workbook_path)
        self.workbook = load_workbook(self.workbook_path)

    @staticmethod
    def copy(file1, file2):
        import shutil
        shutil.copy2(file1, file2)

    @staticmethod
    def str_to_float(string) -> float:
        return float(string.replace(".", "").replace(",", "."))

    @staticmethod
    def return_last_empty_row(spreadsheet):
            for row in spreadsheet.iter_rows(min_row=3, min_col=2, max_col=2):
                if row[0].value is None:
                    return row[0].row

    def write_to_excel(self, spreadsheet, row, page, analysis):
        try:
            spreadsheet.cell(row, 2).value = page.page_number + 1
            spreadsheet.cell(row, 3).value = page.date
            spreadsheet.cell(row, 4).value = page.representative
            spreadsheet.cell(row, 5).value = page.buyer
            spreadsheet.cell(row, 6).value = page.buyer_cpf
            spreadsheet.cell(row, 7).value = page.serial_number
            spreadsheet.cell(row, 8).value = page.analysis_type
            spreadsheet.cell(row, 9).value = self.str_to_float(analysis["kg"])
            spreadsheet.cell(row, 10).value = self.str_to_float(analysis["pd"])
            spreadsheet.cell(row, 11).value = self.str_to_float(analysis["pt"])
            spreadsheet.cell(row, 12).value = self.str_to_float(analysis["rh"])
            spreadsheet.cell(row, 13).value = self.str_to_float(analysis["unitary_value"])
            spreadsheet.cell(row, 14).value = self.str_to_float(analysis["total_value"])
            spreadsheet.cell(row, 15).value = page.quotations[0]
            spreadsheet.cell(row, 16).value = page.quotations[1]
            spreadsheet.cell(row, 17).value = page.quotations[2]

        except TypeError as E:
            spreadsheet.cell(row, 1).value = "error"
            print(row, E)

        except Exception as E:
            print(row, E)
            print(page.analysis_average)

        finally:
            self.workbook.save(self.workbook_path)

    def write(self, folders_list: list, resumed: bool = False):
        spreadsheet = self.workbook.active

        for pdf in folders_list:
            for page in pdf.pages:
                last_empty_row = self.return_last_empty_row(spreadsheet)

                if resumed:
                    self.write_to_excel(spreadsheet, last_empty_row, page, page.analysis_average)

                if not resumed:
                    for i, analysis in enumerate(page.analytics):
                        row = last_empty_row + i
                        self.write_to_excel(spreadsheet, row, page, analysis)


if __name__ == "__main__":
    import pdf

    pdfs = pdf.load()
    wb = Workbook("resumo.xlsx")
    wb.write(pdfs, resumed=False)
    print("done")
